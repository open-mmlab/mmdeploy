# Copyright (c) OpenMMLab. All rights reserved.
import argparse
import glob
import logging
import os
import subprocess
from datetime import datetime
from pathlib import Path
from typing import List, Union

import mmengine
import openpyxl
import pandas as pd
import yaml
from torch.hub import download_url_to_file
from torch.multiprocessing import set_start_method
from tqdm import tqdm

import mmdeploy.version
from mmdeploy.utils import (get_backend, get_codebase, get_root_logger,
                            is_dynamic_shape)


def parse_args():
    parser = argparse.ArgumentParser(description='Regression Test')
    parser.add_argument(
        '--codebase',
        nargs='+',
        help='regression test yaml path.',
        default=[
            'mmpretrain', 'mmdet', 'mmseg', 'mmpose', 'mmocr', 'mmagic',
            'mmrotate', 'mmdet3d'
        ])
    parser.add_argument(
        '-p',
        '--performance',
        default=False,
        action='store_true',
        help='test performance if it set')
    parser.add_argument(
        '--backends', nargs='+', help='test specific backend(s)')
    parser.add_argument(
        '--models', nargs='+', default=['all'], help='test specific model(s)')
    parser.add_argument(
        '--work-dir',
        type=str,
        help='the dir to save logs and models',
        default='../mmdeploy_regression_working_dir')
    parser.add_argument(
        '--checkpoint-dir',
        type=str,
        help='the dir to save checkpoint for all model',
        default='../mmdeploy_checkpoints')
    parser.add_argument(
        '--device',
        type=str,
        help='Device type, cuda:id or cpu, cuda:0 as default',
        default='cuda:0')
    parser.add_argument(
        '--log-level',
        help='set log level',
        default='INFO',
        choices=list(logging._nameToLevel.keys()))
    args = parser.parse_args()

    return args


def merge_report(work_dir: str, logger: logging.Logger):
    """Merge all the report into one report.

    Args:
        work_dir (str): Work dir that including all reports.
        logger (logging.Logger): Logger handler.
    """
    work_dir = Path(work_dir)
    res_file = work_dir.joinpath(
        f'mmdeploy_regression_test_{mmdeploy.version.__version__}.xlsx')
    logger.info(f'Whole result report saving in {res_file}')
    if res_file.exists():
        # delete if it existed
        res_file.unlink()
    for report_file in work_dir.iterdir():
        if report_file.name.startswith('.~'):
            # skip unclosed temp file
            continue
        if '_report.xlsx' not in report_file.name or \
                report_file.name == res_file.name or \
                not report_file.is_file():
            # skip other file
            continue
        # get info from report
        logger.info(f'Merging {report_file}')
        df = pd.read_excel(str(report_file))
        df.rename(columns={'Unnamed: 0': 'Index'}, inplace=True)

        # get key then convert to list
        keys = list(df.keys())
        values = df.values.tolist()

        # sheet name
        sheet_name = report_file.stem.split('_')[0]

        # begin to write
        if res_file.exists():
            # load if it existed
            wb = openpyxl.load_workbook(str(res_file))
        else:
            wb = openpyxl.Workbook()

        # delete if sheet already exist
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        # create sheet
        wb.create_sheet(title=sheet_name, index=0)
        # write in row
        wb[sheet_name].append(keys)
        for value in values:
            wb[sheet_name].append(value)
        # delete the blank sheet
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.cell(1, 1).value is None:
                wb.remove(ws)
        # save to file
        wb.save(str(res_file))

    logger.info('Report merge successful.')


def get_model_metafile_info(global_info: dict, model_info: dict,
                            logger: logging.Logger):
    """Get model metafile information.

    Args:
        global_info (dict): global info from deploy yaml.
        model_info (dict):  model info from deploy yaml.
        logger (logging.Logger): Logger handler.

    Returns:
        Dict: Meta info of each model config
    """

    # get info from global_info and model_info
    checkpoint_dir = global_info.get('checkpoint_dir', None)
    assert checkpoint_dir is not None

    codebase_dir = global_info.get('codebase_dir', None)
    assert codebase_dir is not None

    codebase_name = global_info.get('codebase_name', None)
    assert codebase_name is not None

    model_config_files = model_info.get('model_configs', [])
    assert len(model_config_files) > 0

    # make checkpoint save directory
    model_name = _filter_string(model_info.get('name', 'model'))
    checkpoint_save_dir = Path(checkpoint_dir).joinpath(
        codebase_name, model_name)
    checkpoint_save_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f'Saving checkpoint in {checkpoint_save_dir}')

    # get model metafile info
    metafile_path = Path(codebase_dir).joinpath(model_info.get('metafile'))
    if not metafile_path.exists():
        logger.warning(f'Metafile not exists: {metafile_path}')
        return [], '', ''
    with open(metafile_path) as f:
        metafile_info = yaml.load(f, Loader=yaml.FullLoader)

    model_meta_info = dict()
    for meta_model in metafile_info.get('Models'):
        if str(meta_model.get('Config')) not in model_config_files:
            # skip if the model not in model_config_files
            logger.warning(f'{meta_model.get("Config")} '
                           f'not in {model_config_files}, pls check ! '
                           'Skip it...')
            continue

        # get meta info
        model_meta_info.update({meta_model.get('Config'): meta_model})

        # get weight url
        weights_url = meta_model.get('Weights')
        weights_name = str(weights_url).split('/')[-1]
        weights_save_path = checkpoint_save_dir.joinpath(weights_name)
        if weights_save_path.exists() and \
                not global_info.get('checkpoint_force_download', False):
            logger.info(f'model {weights_name} exist, skip download it...')
            continue

        # Download weight
        logger.info(f'Downloading {weights_url} to {weights_save_path}')
        download_url_to_file(
            weights_url, str(weights_save_path), progress=True)

        # check weather the weight download successful
        if not weights_save_path.exists():
            raise FileExistsError(f'Weight {weights_name} download fail')

    logger.info('All models had been downloaded successful !')
    return model_meta_info, checkpoint_save_dir, codebase_dir


def update_report(report_dict: dict, model_name: str, model_config: str,
                  task_name: str, checkpoint: str, dataset: str,
                  backend_name: str, deploy_config: str,
                  static_or_dynamic: str, precision_type: str,
                  conversion_result: str, fps: str, metric_info: list,
                  test_pass: str, report_txt_path: Path, codebase_name: str):
    """Update report information.

    Args:
        report_dict (dict): Report info dict.
        model_name (str): Model name.
        model_config (str): Model config name.
        task_name (str): Task name.
        checkpoint (str): Model checkpoint name.
        dataset (str): Dataset name.
        backend_name (str): Backend name.
        deploy_config (str): Deploy config name.
        static_or_dynamic (str): Static or dynamic.
        precision_type (str): Precision type of the model.
        conversion_result (str): Conversion result: Successful or Fail.
        fps (str): Inference speed (ms/im).
        metric_info (list): Metric info list of the ${modelName}.yml.
        test_pass (str): Test result: Pass or Fail.
        report_txt_path (Path): Report txt path.
        codebase_name (str): Codebase name.
    """
    # save to tmp file
    tmp_str = f'{model_name},{model_config},{task_name},{checkpoint},' \
              f'{dataset},{backend_name},{deploy_config},' \
              f'{static_or_dynamic},{precision_type},{conversion_result},'

    # save to report
    report_dict.get('Model').append(model_name)
    report_dict.get('Model Config').append(model_config)
    report_dict.get('Task').append(task_name)
    report_dict.get('Checkpoint').append(checkpoint)
    report_dict.get('Dataset').append(dataset)
    report_dict.get('Backend').append(backend_name)
    report_dict.get('Deploy Config').append(deploy_config)
    report_dict.get('Static or Dynamic').append(static_or_dynamic)
    report_dict.get('Precision Type').append(precision_type)
    report_dict.get('Conversion Result').append(conversion_result)
    # report_dict.get('FPS').append(fps)

    for metric in metric_info:
        for metric_name, metric_value in metric.items():
            metric_name = str(metric_name)
            report_dict.get(metric_name).append(metric_value)
            tmp_str += f'{metric_value},'
    report_dict.get('Test Pass').append(test_pass)

    tmp_str += f'{test_pass}\n'

    with open(report_txt_path, 'a+', encoding='utf-8') as f:
        f.write(tmp_str)


def get_pytorch_result(model_name: str, meta_info: dict, checkpoint_path: Path,
                       model_config_path: Path, model_config_name: str,
                       test_yaml_metric_info: dict, report_dict: dict,
                       logger: logging.Logger, report_txt_path: Path,
                       codebase_name: str):
    """Get metric from metafile info of the model.

    Args:
        model_name (str): Name of model.
        meta_info (dict): Metafile info from model's metafile.yml.
        checkpoint_path (Path): Checkpoint path.
        model_config_path (Path): Model config path.
        model_config_name (str): Name of model config in meta_info.
        test_yaml_metric_info (dict): Metrics info from test yaml.
        report_dict (dict): Report info dict.
        logger (logging.Logger): Logger.
        report_txt_path (Path): Report txt path.
        codebase_name (str): Codebase name.

    Returns:
        Dict: metric info of the model
    """

    if model_config_name not in meta_info:
        logger.warning(
            f'{model_config_name} not in meta_info, which is {meta_info}')
        return {}

    # get metric
    model_info = meta_info[model_config_name]
    metafile_metric_info = model_info['Results']
    # deal with mmseg case
    if not isinstance(metafile_metric_info, (list, tuple)):
        metafile_metric_info = [metafile_metric_info]
    pytorch_metric = dict()
    using_dataset = set()
    using_task = set()
    datasets = []

    # Get metrics info from metafile
    for metafile_metric in metafile_metric_info:
        task_name = metafile_metric['Task']
        dataset = metafile_metric['Dataset']

        # check if metafile use the same metric on several datasets for mmagic
        task_info = set([_['Task'] for _ in metafile_metric_info])
        if len(metafile_metric_info) > 1 and len(task_info) == 1:
            for k, v in metafile_metric['Metrics'].items():
                pytorch_metric[f'{dataset} {k}'] = v
        else:
            pytorch_metric.update(metafile_metric['Metrics'])
        datasets.append(dataset)
        using_task.add(task_name)
        using_dataset.add(dataset)

    dataset_type = '+'.join(list(using_dataset))
    task_type = '+'.join(list(using_task))
    metric_list = []
    for metric, metric_info in test_yaml_metric_info.items():
        value = '-'
        if metric in pytorch_metric:
            if 'dataset' in metric_info and metric_info['dataset'] in datasets:
                idx = datasets.index(metric_info['dataset'])
                pytorch_metric.update(metafile_metric_info[idx]['Metrics'])
            value = pytorch_metric[metric]
        metric_list.append({metric: value})

    valid_pytorch_metric = {
        k: v
        for k, v in pytorch_metric.items() if k in test_yaml_metric_info
    }

    # get pytorch fps value
    fps_info = model_info.get('Metadata', {}).get('inference time (ms/im)')
    if fps_info is None:
        fps = '-'
    elif isinstance(fps_info, list):
        fps = fps_info[0].get('value')
    else:
        fps = fps_info.get('value')

    logger.info(f'Got metric_list = {metric_list} ')
    logger.info(f'Got pytorch_metric = {pytorch_metric} ')

    # update report
    update_report(
        report_dict=report_dict,
        model_name=model_name,
        model_config=str(model_config_path),
        task_name=task_type,
        checkpoint=str(checkpoint_path),
        dataset=dataset_type,
        backend_name='Pytorch',
        deploy_config='-',
        static_or_dynamic='-',
        precision_type='-',
        conversion_result='-',
        fps=fps,
        metric_info=metric_list,
        test_pass='-',
        report_txt_path=report_txt_path,
        codebase_name=codebase_name)

    logger.info(f'Got {model_config_path} metric: {valid_pytorch_metric}')
    dataset_info = dict(dataset=dataset_type, task=task_type)
    return valid_pytorch_metric, dataset_info


def parse_test_log(work_dir: str) -> dict:
    """Parse metrics result from output json file.

    Args:
        work_dir: work directory that has output json file.

    Returns:
        dict: metric results
    """
    logger = get_root_logger()
    json_files = glob.glob(os.path.join(work_dir, '*', '*.json'))
    json_path = None
    newest_date = None
    # filter json and get latest json file
    for f in json_files:
        fname = os.path.split(f)[1].strip('.json')
        try:
            date = datetime.strptime(fname, '%Y%m%d_%H%M%S')
            if newest_date is None:
                newest_date = date
                json_path = f
            elif date > newest_date:
                newest_date = date
                json_path = f
        except Exception:
            pass
    if (not os.path.exists(work_dir)) or json_path is None:
        logger.warning(f'Not json files found in {work_dir}')
        result = {}
    else:
        logger.info(f'Parse test result from {json_path}')
        result = mmengine.load(json_path)
    return result


def get_fps_metric(shell_res: int, pytorch_metric: dict, metric_info: dict,
                   work_path: Path):
    """Get fps and metric.

    Args:
        shell_res (int): Backend convert result: 0 is success.
        pytorch_metric (dict): Metric info of pytorch metafile.
        work_path (Path): Logger path.
        metric_info (dict): Metric info.

    Returns:
        Float: fps: FPS of the model.
        List: metric_list: metric result list.
        Bool: test_pass: If the test pass or not.
    """
    # check if converted successes or not.
    fps = '-'
    if shell_res != 0:
        backend_results = {}
    else:
        backend_results = parse_test_log(work_path)
    compare_results = {}
    output_result = {}

    for metric_name, metric_value in pytorch_metric.items():
        metric_key = metric_info[metric_name]['metric_key']
        tolerance = metric_info[metric_name]['tolerance']
        multi_value = metric_info[metric_name].get('multi_value', 1.0)
        compare_flag = False
        output_result[metric_name] = 'x'
        if metric_key in backend_results:
            backend_value = backend_results[metric_key] * multi_value
            output_result[metric_name] = backend_value
            if backend_value >= metric_value - tolerance:
                compare_flag = True
        compare_results[metric_name] = compare_flag

    if len(compare_results):
        test_pass = all(list(compare_results.values()))
    else:
        test_pass = False
    return fps, output_result, test_pass


def get_backend_fps_metric(deploy_cfg_path: str, model_cfg_path: Path,
                           convert_checkpoint_path: str, device_type: str,
                           logger: logging.Logger, pytorch_metric: dict,
                           metric_info: dict, backend_name: str,
                           precision_type: str, convert_result: bool,
                           report_dict: dict, infer_type: str, log_path: Path,
                           dataset_info: dict, report_txt_path: Path,
                           model_name: str):
    """Get backend fps and metric.

    Args:
        deploy_cfg_path (str): Deploy config path.
        model_cfg_path (Path): Model config path.
        convert_checkpoint_path (str): Converted checkpoint path.
        device_type (str): Device for converting.
        logger (logging.Logger): Logger handler.
        pytorch_metric (dict): Pytorch metric info dict get from metafile.
        metric_info (dict): Metric info from test yaml.
        backend_name (str): Backend name.
        precision_type (str): Precision type for evaluation.
        convert_result (bool): Backend convert result.
        report_dict (dict): Backend convert result.
        infer_type (str): Infer type.
        log_path (Path): Logger save path.
        dataset_info (dict): Dataset info.
        report_txt_path (Path): report txt save path.
        model_name (str): Name of model in test yaml.
    """
    work_dir = log_path.parent.joinpath('test_logs')
    if not work_dir.exists():
        work_dir.mkdir(parents=True, exist_ok=True)
    cmd_lines = [
        'python3 tools/test.py', f'{deploy_cfg_path}', f'{model_cfg_path}',
        f'--model {convert_checkpoint_path}', f'--work-dir "{work_dir}"',
        '--speed-test', f'--device {device_type}'
    ]

    codebase_name = get_codebase(str(deploy_cfg_path)).value
    # to stop Dataloader OOM in docker CI
    if codebase_name not in ['mmagic', 'mmocr', 'mmpretrain']:
        cfg_options = 'test_dataloader.num_workers=0 ' \
                      'test_dataloader.persistent_workers=False ' \
                      'val_dataloader.num_workers=0 ' \
                      'val_dataloader.persistent_workers=False '
        cmd_lines.append(f'--cfg-options {cfg_options}')

    # Test backend
    return_code = run_cmd(cmd_lines, log_path)
    fps, backend_metric, test_pass = get_fps_metric(return_code,
                                                    pytorch_metric,
                                                    metric_info, work_dir)
    logger.info(f'test_pass= {test_pass}, results= {backend_metric}')
    metric_list = []
    for metric in metric_info:
        value = '-'
        if metric in backend_metric:
            value = backend_metric[metric]
        metric_list.append({metric: value})
    dataset_type = dataset_info['dataset']
    task_name = dataset_info['task']
    # update the report
    update_report(
        report_dict=report_dict,
        model_name=model_name,
        model_config=str(model_cfg_path),
        task_name=task_name,
        checkpoint=convert_checkpoint_path,
        dataset=dataset_type,
        backend_name=backend_name,
        deploy_config=str(deploy_cfg_path),
        static_or_dynamic=infer_type,
        precision_type=precision_type,
        conversion_result=str(convert_result),
        fps=fps,
        metric_info=metric_list,
        test_pass=str(test_pass),
        report_txt_path=report_txt_path,
        codebase_name=codebase_name)


def get_precision_type(deploy_cfg_name: str):
    """Get backend precision_type according to the name of deploy config.

    Args:
        deploy_cfg_name (str): Name of the deploy config.

    Returns:
        Str: precision_type: Precision type of the deployment name.
    """
    if 'int8' in deploy_cfg_name:
        precision_type = 'int8'
    elif 'fp16' in deploy_cfg_name:
        precision_type = 'fp16'
    else:
        precision_type = 'fp32'

    return precision_type


def replace_top_in_pipeline_json(backend_output_path: Path,
                                 logger: logging.Logger):
    """Replace `topk` with `num_classes` in `pipeline.json`.

    Args:
        backend_output_path (Path): Backend convert result path.
        logger (logger.Logger): Logger handler.
    """

    sdk_pipeline_json_path = backend_output_path.joinpath('pipeline.json')
    sdk_pipeline_json = mmengine.load(sdk_pipeline_json_path)

    pipeline_tasks = sdk_pipeline_json.get('pipeline', {}).get('tasks', [])
    for index, task in enumerate(pipeline_tasks):
        if task.get('name', '') != 'postprocess':
            continue
        num_classes = task.get('params', {}).get('num_classes', 0)
        if 'topk' not in task.get('params', {}):
            continue
        sdk_pipeline_json['pipeline']['tasks'][index]['params']['topk'] = \
            num_classes

    logger.info(f'sdk_pipeline_json = {sdk_pipeline_json}')

    mmengine.dump(
        sdk_pipeline_json, sdk_pipeline_json_path, sort_keys=False, indent=4)

    logger.info('replace done')


def gen_log_path(backend_output_path: Path, log_name: str):
    if not backend_output_path.exists():
        backend_output_path.mkdir(parents=True, exist_ok=True)
    log_path = backend_output_path.joinpath(log_name)
    if log_path.exists():
        os.remove(str(log_path))
    return log_path


def run_cmd(cmd_lines: List[str], log_path: Path):
    """
    Args:
        cmd_lines: (list[str]): A command in multiple line style.
        log_path (Path): Path to log file.

    Returns:
        int: error code.
    """
    import platform
    system = platform.system().lower()

    if system == 'windows':
        sep = r'`'
    else:  # 'Linux', 'Darwin'
        sep = '\\'
    cmd_for_run = ' '.join(cmd_lines)
    cmd_for_log = f' {sep}\n'.join(cmd_lines) + '\n'
    parent_path = log_path.parent
    if not parent_path.exists():
        parent_path.mkdir(parents=True, exist_ok=True)
    logger = get_root_logger()
    logger.info(100 * '-')
    logger.info(f'Start running cmd\n{cmd_for_log}')
    logger.info(f'Logging log to \n{log_path}')

    with open(log_path, 'w', encoding='utf-8') as file_handler:
        # write cmd
        file_handler.write(f'Command:\n{cmd_for_log}\n')
        file_handler.flush()
        process_res = subprocess.Popen(
            cmd_for_run,
            cwd=str(Path(__file__).absolute().parent.parent),
            shell=True,
            stdout=file_handler,
            stderr=file_handler)
        process_res.wait()
        return_code = process_res.returncode

    if return_code != 0:
        logger.error(f'Got shell return code={return_code}')
        with open(log_path, 'r') as f:
            content = f.read()
            logger.error(f'Log error message\n{content}')
    return return_code


def get_backend_result(pipeline_info: dict, model_cfg_path: Path,
                       checkpoint_path: Path, work_dir: Path, device_type: str,
                       pytorch_metric: dict, metric_info: dict,
                       report_dict: dict, test_type: str,
                       logger: logging.Logger, backend_file_name: Union[str,
                                                                        list],
                       report_txt_path: Path, metafile_dataset: str,
                       model_name: str):
    """Convert model to onnx and then get metric.

    Args:
        pipeline_info (dict):  Pipeline info of test yaml.
        model_cfg_path (Path): Model config file path.
        checkpoint_path (Path): Checkpoints path.
        work_dir (Path): A working directory.
        device_type (str): A string specifying device, defaults to 'cuda'.
        pytorch_metric (dict): All pytorch metric info.
        metric_info (dict): Metrics info.
        report_dict (dict): Report info dict.
        test_type (str): Test type. 'precision' or 'convert'.
        logger (logging.Logger): Logger.
        backend_file_name (str | list): backend file save name.
        report_txt_path (Path): report txt path.
        metafile_dataset (str): Dataset type get from metafile.
        model_name (str): Name of model in test yaml.
    """
    # get backend_test info
    backend_test = pipeline_info.get('backend_test', False)

    # get convert_image info
    convert_image_info = pipeline_info.get('convert_image', None)
    if convert_image_info is not None:
        input_img_path = \
            convert_image_info.get('input_img', './tests/data/tiger.jpeg')
        test_img_path = convert_image_info.get('test_img', None)
    else:
        input_img_path = './tests/data/tiger.jpeg'
        test_img_path = None
    # get sdk_cfg info
    sdk_config = pipeline_info.get('sdk_config', None)
    if sdk_config is not None:
        sdk_config = Path(sdk_config)

    # Overwrite metric tolerance
    metric_tolerance = pipeline_info.get('metric_tolerance', None)
    if metric_tolerance is not None:
        for metric, new_tolerance in metric_tolerance.items():
            if metric not in metric_info:
                logger.debug(f'{metric} not in {metric_info},'
                             'skip compare it...')
                continue
            if new_tolerance is None:
                logger.debug('new_tolerance is None, skip it ...')
                continue
            metric_info[metric]['tolerance'] = new_tolerance
    if backend_test is False and sdk_config is None:
        test_type = 'convert'

    deploy_cfg_path = Path(pipeline_info.get('deploy_config'))
    backend_name = str(get_backend(str(deploy_cfg_path)).name).lower()

    # change device_type for special case
    if backend_name in ['ncnn', 'openvino']:
        device_type = 'cpu'
    elif backend_name == 'onnxruntime' and device_type != 'cpu':
        import onnxruntime as ort
        if ort.get_device() != 'GPU':
            device_type = 'cpu'
            logger.warning('Device type is forced to cpu '
                           'since onnxruntime-gpu is not installed')

    infer_type = \
        'dynamic' if is_dynamic_shape(str(deploy_cfg_path)) else 'static'

    precision_type = get_precision_type(deploy_cfg_path.name)
    codebase_name = get_codebase(str(deploy_cfg_path)).value

    backend_output_path = Path(work_dir). \
        joinpath(Path(checkpoint_path).parent.parent.name,
                 Path(checkpoint_path).parent.name,
                 backend_name,
                 infer_type,
                 precision_type,
                 Path(checkpoint_path).stem)
    backend_output_path.mkdir(parents=True, exist_ok=True)

    # convert cmd lines
    cmd_lines = [
        'python3 ./tools/deploy.py', f'{deploy_cfg_path}', f'{model_cfg_path}',
        f'"{checkpoint_path}"', f'"{input_img_path}"',
        f'--work-dir "{backend_output_path}"', f'--device {device_type}',
        '--log-level INFO'
    ]

    if sdk_config is not None and test_type == 'precision':
        cmd_lines += ['--dump-info']

    if test_img_path is not None:
        cmd_lines += [f'--test-img {test_img_path}']

    if precision_type == 'int8':
        calib_dataset_cfg = pipeline_info.get('calib_dataset_cfg', None)
        if calib_dataset_cfg is not None:
            cmd_lines += [f'--calib-dataset-cfg {calib_dataset_cfg}']

    convert_log_path = backend_output_path.joinpath('convert_log.txt')
    return_code = run_cmd(cmd_lines, convert_log_path)
    convert_result = return_code == 0
    logger.info(f'Got convert_result = {convert_result}')

    if isinstance(backend_file_name, list):
        report_checkpoint = backend_output_path.joinpath(backend_file_name[0])
        convert_checkpoint_path = ''
        for backend_file in backend_file_name:
            backend_path = backend_output_path.joinpath(backend_file)
            convert_checkpoint_path += f'{backend_path} '
    else:
        report_checkpoint = backend_output_path.joinpath(backend_file_name)
        convert_checkpoint_path = \
            str(backend_output_path.joinpath(backend_file_name))

    # Test the model
    if convert_result and test_type == 'precision':
        # test the model metric
        if backend_test:
            log_path = \
                gen_log_path(backend_output_path.joinpath('backend'),
                             'test_log.txt')

            get_backend_fps_metric(
                deploy_cfg_path=str(deploy_cfg_path),
                model_cfg_path=model_cfg_path,
                convert_checkpoint_path=convert_checkpoint_path,
                device_type=device_type,
                logger=logger,
                pytorch_metric=pytorch_metric,
                metric_info=metric_info,
                backend_name=backend_name,
                precision_type=precision_type,
                convert_result=convert_result,
                report_dict=report_dict,
                infer_type=infer_type,
                log_path=log_path,
                dataset_info=metafile_dataset,
                report_txt_path=report_txt_path,
                model_name=model_name)

        if sdk_config is not None:

            if codebase_name == 'mmpretrain' or codebase_name == 'mmaction':
                replace_top_in_pipeline_json(backend_output_path, logger)

            log_path = gen_log_path(
                backend_output_path.joinpath('sdk'), 'test_log.txt')
            if backend_name == 'onnxruntime':
                # sdk only support onnxruntime of cpu
                device_type = 'cpu'
            # sdk test
            get_backend_fps_metric(
                deploy_cfg_path=str(sdk_config),
                model_cfg_path=model_cfg_path,
                convert_checkpoint_path=str(backend_output_path),
                device_type=device_type,
                logger=logger,
                pytorch_metric=pytorch_metric,
                metric_info=metric_info,
                backend_name=f'SDK-{backend_name}',
                precision_type=precision_type,
                convert_result=convert_result,
                report_dict=report_dict,
                infer_type=infer_type,
                log_path=log_path,
                dataset_info=metafile_dataset,
                report_txt_path=report_txt_path,
                model_name=model_name)
    else:
        logger.info('Only test convert, saving to report...')
        metric_list = [{metric: '-'} for metric in metric_info]
        fps = '-'
        test_pass = convert_result
        dataset_type = metafile_dataset['dataset']
        task_name = metafile_dataset['task']
        # update the report
        update_report(
            report_dict=report_dict,
            model_name=model_name,
            model_config=str(model_cfg_path),
            task_name=task_name,
            checkpoint=str(report_checkpoint),
            dataset=dataset_type,
            backend_name=backend_name,
            deploy_config=str(deploy_cfg_path),
            static_or_dynamic=infer_type,
            precision_type=precision_type,
            conversion_result=str(convert_result),
            fps=fps,
            metric_info=metric_list,
            test_pass=str(test_pass),
            report_txt_path=report_txt_path,
            codebase_name=codebase_name)


def save_report(report_info: dict, report_save_path: Path,
                logger: logging.Logger):
    """Convert model to onnx and then get metric.

    Args:
        report_info (dict):  Report info dict.
        report_save_path (Path): Report save path.
        logger (logging.Logger): Logger.
    """
    logger.info('Saving regression test report to '
                f'{report_save_path}, pls wait...')
    try:
        df = pd.DataFrame(report_info)
        df.to_excel(report_save_path)
    except ValueError:
        logger.info(f'Got error report_info = {report_info}')

    logger.info('Saved regression test report to '
                f'{report_save_path}.')


def _filter_string(inputs):
    """Remove non alpha&number character from input string.

    Args:
        inputs(str): Input string.

    Returns:
        str: Output of only alpha&number string.
    """
    outputs = ''.join([i.lower() for i in inputs if i.isalnum()])
    return outputs


def main():
    args = parse_args()
    set_start_method('spawn')
    logger = get_root_logger(log_level=args.log_level)

    test_type = 'precision' if args.performance else 'convert'
    logger.info(f'Processing regression test in {test_type} mode.')

    backend_file_info = {
        'onnxruntime': 'end2end.onnx',
        'tensorrt': 'end2end.engine',
        'openvino': 'end2end.xml',
        'ncnn': ['end2end.param', 'end2end.bin'],
        'pplnn': ['end2end.onnx', 'end2end.json'],
        'torchscript': 'end2end.pt'
    }

    backend_list = args.backends
    if backend_list is None:
        backend_list = [
            'onnxruntime', 'tensorrt', 'openvino', 'ncnn', 'pplnn',
            'torchscript'
        ]
    assert isinstance(backend_list, list)
    logger.info(f'Regression test backend list = {backend_list}')

    args.models = [_filter_string(s) for s in args.models]
    logger.info(f'Regression test models list = {args.models}')

    assert ' ' not in args.work_dir, \
        f'No empty space included in {args.work_dir}'
    assert ' ' not in args.checkpoint_dir, \
        f'No empty space included in {args.checkpoint_dir}'

    work_dir = Path(args.work_dir)
    work_dir.mkdir(parents=True, exist_ok=True)

    deploy_yaml_list = [
        f'./tests/regression/{codebase}.yml' for codebase in args.codebase
    ]

    for deploy_yaml in deploy_yaml_list:
        if not Path(deploy_yaml).exists():
            raise FileNotFoundError(f'deploy_yaml {deploy_yaml} not found, '
                                    'please check !')

        with open(deploy_yaml) as f:
            yaml_info = yaml.load(f, Loader=yaml.FullLoader)

        report_save_path = \
            work_dir.joinpath(Path(deploy_yaml).stem + '_report.xlsx')
        report_txt_path = report_save_path.with_suffix('.txt')

        report_dict = {
            'Model': [],
            'Model Config': [],
            'Task': [],
            'Checkpoint': [],
            'Dataset': [],
            'Backend': [],
            'Deploy Config': [],
            'Static or Dynamic': [],
            'Precision Type': [],
            'Conversion Result': [],
            # 'FPS': []
        }

        global_info = yaml_info.get('globals')
        metric_info = global_info.get('metric_info', {})
        for metric_name in metric_info:
            report_dict.update({metric_name: []})
        report_dict.update({'Test Pass': []})

        global_info.update({'checkpoint_dir': args.checkpoint_dir})
        global_info.update(
            {'codebase_name': Path(deploy_yaml).stem.split('_')[0]})

        with open(report_txt_path, 'w') as f_report:
            title_str = ''
            for key in report_dict:
                title_str += f'{key},'
            title_str = title_str[:-1] + '\n'
            f_report.write(title_str)  # clear the report tmp file

        models_info = yaml_info.get('models')
        for models in tqdm(models_info):
            model_name_origin = models.get('name', 'model')
            model_name_new = _filter_string(model_name_origin)
            if 'model_configs' not in models:
                logger.warning('Can not find field "model_configs", '
                               f'skipping {model_name_origin}...')
                continue

            if args.models != ['all'] and model_name_new not in args.models:
                logger.info(
                    f'Test specific model mode, skip {model_name_origin}...')
                continue
            try:
                model_metafile_info, checkpoint_save_dir, codebase_dir = \
                    get_model_metafile_info(global_info, models, logger)
            except Exception as e:
                logger.error(f'Failed to get meta info {e}')
                continue
            for model_config in model_metafile_info:
                logger.info(f'Processing test for {model_config}...')

                # Get backends info
                pipelines_info = models.get('pipelines', None)
                if pipelines_info is None:
                    logger.warning('pipelines_info is None, skip it...')
                    continue

                # Get model config path
                model_cfg_path = Path(codebase_dir).joinpath(model_config)
                assert model_cfg_path.exists()

                # Get checkpoint path
                checkpoint_name = Path(
                    model_metafile_info.get(model_config).get('Weights')).name

                checkpoint_path = Path(checkpoint_save_dir, checkpoint_name)
                assert checkpoint_path.exists()

                # Get pytorch from metafile.yml
                pytorch_metric, metafile_dataset = get_pytorch_result(
                    model_name_origin, model_metafile_info, checkpoint_path,
                    model_cfg_path, model_config, metric_info, report_dict,
                    logger, report_txt_path, global_info.get('codebase_name'))
                for pipeline in pipelines_info:
                    deploy_config = pipeline.get('deploy_config')
                    backend_name = get_backend(deploy_config).name.lower()
                    if backend_name not in backend_list:
                        logger.warning(f'backend_name ({backend_name}) not '
                                       f'in {backend_list}, skip it...')
                        continue

                    backend_file_name = \
                        backend_file_info.get(backend_name, None)
                    if backend_file_name is None:
                        logger.warning('backend_file_name is None, '
                                       'skip it...')
                        continue

                    get_backend_result(pipeline, model_cfg_path,
                                       checkpoint_path, work_dir, args.device,
                                       pytorch_metric, metric_info,
                                       report_dict, test_type, logger,
                                       backend_file_name, report_txt_path,
                                       metafile_dataset, model_name_origin)
        if len(report_dict.get('Model')) > 0:
            save_report(report_dict, report_save_path, logger)
        else:
            logger.info(f'No model for {deploy_yaml}, not saving report.')

    # merge report
    merge_report(str(work_dir), logger)

    logger.info('All done.')


if __name__ == '__main__':
    main()
